#!/usr/bin/env python3
"""
股票板块轮动跟踪器 - 本地服务器
支持静态文件 + Excel导出API
"""
import http.server
import json
import os
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')

class StockHandler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        
        if parsed.path == '/export_excel':
            self.handle_export_excel(parsed)
        else:
            super().do_GET()
    
    def handle_export_excel(self, parsed):
        """生成并返回Excel文件"""
        params = parse_qs(parsed.query)
        date = params.get('date', [datetime.now().strftime('%Y-%m-%d')])[0]
        
        # 读取数据
        data_file = os.path.join(DATA_DIR, f'{date.replace("-", "")}.json')
        if not os.path.exists(data_file):
            data_file = os.path.join(DATA_DIR, 'latest.json')
        
        if not os.path.exists(data_file):
            self.send_error(404, '数据文件不存在')
            return
        
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 生成Excel
        wb = Workbook()
        
        # 样式定义
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        red_font = Font(color='FF0000')
        green_font = Font(color='008000')
        number_align = Alignment(horizontal='right')
        
        def setup_sheet(ws, title, headers, col_widths):
            ws.title = title
            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        
        def add_row(ws, row_idx, values, color_col=None):
            for i, v in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=i, value=v)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center') if i == 1 else number_align
                if color_col is not None and i == color_col:
                    if isinstance(v, (int, float)):
                        cell.font = red_font if v > 0 else green_font if v < 0 else Font()
        
        # === Sheet 1: 涨停股 ===
        ws1 = wb.active
        headers1 = ['股票代码', '股票名称', '行业板块', '涨跌幅(%)', '成交额(万)', '换手率(%)', '连板数', '首次封板', '封单金额(万)']
        widths1 = [12, 14, 14, 12, 14, 12, 10, 12, 14]
        setup_sheet(ws1, '涨停股', headers1, widths1)
        
        for i, s in enumerate(data.get('zt_stocks', []), 2):
            vals = [
                s.get('code', ''),
                s.get('name', ''),
                s.get('industry', ''),
                round(s.get('change_pct', 0), 2),
                round(s.get('amount', 0) / 10000, 0),
                round(s.get('turnover', 0), 2),
                s.get('continuous', 1),
                s.get('first_time', ''),
                round(s.get('seal_amount', 0) / 10000, 0) if s.get('seal_amount') else '',
            ]
            add_row(ws1, i, vals, color_col=4)
        
        # === Sheet 2: 行业板块资金流向 ===
        ws2 = wb.create_sheet()
        headers2 = ['板块名称', '主力净流入(亿)', '涨跌幅(%)', '上涨家数', '下跌家数']
        widths2 = [16, 16, 12, 10, 10]
        setup_sheet(ws2, '行业板块资金流向', headers2, widths2)
        
        for i, s in enumerate(data.get('industry_flow', []), 2):
            vals = [
                s.get('name', ''),
                round(s.get('net_inflow', 0) / 1e8, 2),
                round(s.get('change_pct', 0), 2),
                s.get('rise_count', 0),
                s.get('fall_count', 0),
            ]
            add_row(ws2, i, vals, color_col=2)
        
        # === Sheet 3: 概念板块涨幅 ===
        ws3 = wb.create_sheet()
        headers3 = ['概念名称', '涨跌幅(%)', '主力净流入(亿)']
        widths3 = [18, 12, 16]
        setup_sheet(ws3, '概念板块涨幅', headers3, widths3)
        
        for i, s in enumerate(data.get('concept_flow', []), 2):
            vals = [
                s.get('name', ''),
                round(s.get('change_pct', 0), 2),
                round(s.get('net_inflow', 0) / 1e8, 2),
            ]
            add_row(ws3, i, vals, color_col=2)
        
        # === Sheet 4: 板块成交量 ===
        ws4 = wb.create_sheet()
        headers4 = ['板块名称', '成交额(亿)', '涨跌幅(%)', '换手率(%)']
        widths4 = [16, 14, 12, 12]
        setup_sheet(ws4, '板块成交量', headers4, widths4)
        
        for i, s in enumerate(data.get('sector_volume', []), 2):
            vals = [
                s.get('name', ''),
                round(s.get('volume', 0) / 1e8, 2),
                round(s.get('change_pct', 0), 2),
                round(s.get('turnover', 0), 2),
            ]
            add_row(ws4, i, vals, color_col=3)
        
        # === Sheet 5: 大盘数据 ===
        ws5 = wb.create_sheet()
        headers5 = ['日期', '收盘价', '涨跌幅(%)', '成交额(亿)']
        widths5 = [14, 12, 12, 14]
        setup_sheet(ws5, '大盘数据', headers5, widths5)
        
        mkt = data.get('market', {})
        if mkt:
            add_row(ws5, 2, [
                mkt.get('date', ''),
                round(mkt.get('close', 0), 2),
                round(mkt.get('change_pct', 0), 2),
                round(mkt.get('volume', 0) / 1e8, 0),
            ], color_col=3)
        
        # 输出到内存
        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        # 发送响应
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', f'attachment; filename="stock_board_{date}.xlsx"')
        self.send_header('Content-Length', str(len(buf.getvalue())))
        self.end_headers()
        self.wfile.write(buf.getvalue())
    
    def log_message(self, format, *args):
        if '/export_excel' not in str(args):
            super().log_message(format, *args)


if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    server = http.server.HTTPServer(('0.0.0.0', 18080), StockHandler)
    print('服务器启动: http://localhost:18080')
    server.serve_forever()
